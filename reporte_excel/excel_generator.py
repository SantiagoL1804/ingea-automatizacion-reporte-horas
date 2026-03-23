#!/usr/bin/env python3
import requests
import os
from datetime import datetime, timedelta
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from typing import List, Dict, Optional
import re

# Load environment variables
load_dotenv()


class ClockifyToExcelConverter:
    """Convierte registros de Clockify a una planilla Excel con formato específico"""
    
    def __init__(self):
        self.clockify_api_key = os.getenv('CLOCKIFY_API_KEY')
        self.workspace_id = os.getenv('CLOCKIFY_WORKSPACE_ID')
        self.user_id = os.getenv('CLOCKIFY_USER_ID')
        self.project_id = os.getenv('CLOCKIFY_PROJECT_ID')
        self.jira_api_key = os.getenv('JIRA_API_KEY')
        self.jira_url = os.getenv('JIRA_URL', 'https://your-domain.atlassian.net')
        self.jira_email = os.getenv('JIRA_EMAIL')
        
        self.headers_clockify = {
            "X-Api-Key": self.clockify_api_key,
            "Content-Type": "application/json"
        }
    
    def fetch_time_entries(self, start_date: datetime, end_date: datetime) -> List[Dict]:
        """Obtiene los registros de tiempo de Clockify en un rango de fechas"""
        url = f"https://api.clockify.me/api/v1/workspaces/{self.workspace_id}/user/{self.user_id}/time-entries"
        
        params = {
            "start": start_date.isoformat() + "Z",
            "end": end_date.isoformat() + "Z",
            "page-size": 500
        }
        
        if self.project_id:
            params["project"] = self.project_id
        
        try:
            response = requests.get(url, headers=self.headers_clockify, params=params, timeout=30)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            print(f"❌ Error fetching Clockify entries: {e}")
            return []
    
    def search_jira_issue_id(self, description: str) -> tuple[str, str]:
        """Extrae el ID de Jira y limpia la descripción (IAPP-XXX o IAPP XXX)"""
        # Busca patrones como "IAPP-631" o "IAPP 631"
        match = re.search(r'IAPP[-\s]?(\d+)', description, re.IGNORECASE)
        if match:
            # Normaliza al formato IAPP-XXX
            jira_id = f"IAPP-{match.group(1)}"
            cleaned = re.sub(r'\bIAPP[-\s]?\d+\b', '', description, flags=re.IGNORECASE)
            cleaned = ' '.join(cleaned.split()).strip()
            print(f"✓ Found: {jira_id} in '{description[:50]}...'")
            return jira_id, cleaned
        print(f"✗ Not found in: '{description[:50]}...'")
        return "N/A", description
    
    def get_work_location(self, date: datetime) -> str:
        """Determina el lugar de trabajo según el día de la semana"""
        weekday = date.weekday()  # 0=Monday, 6=Sunday
        
        # Miércoles (2), Viernes (4), Sábado (5), Domingo (6) -> Remoto
        if weekday in [2, 4, 5, 6]:
            return "Remoto"
        else:
            return "Oficina"
    
    def calculate_hours(self, time_interval: Dict) -> float:
        """Calcula las horas trabajadas desde el intervalo de tiempo"""
        start_str = time_interval.get('start')
        end_str = time_interval.get('end')
        
        if not start_str or not end_str:
            return 0.0
        
        try:
            start = datetime.fromisoformat(start_str.replace('Z', '+00:00'))
            end = datetime.fromisoformat(end_str.replace('Z', '+00:00'))
            duration = end - start
            hours = duration.total_seconds() / 3600
            # Redondear hacia arriba a múltiplos de 0.5
            import math
            return math.ceil(hours * 2) / 2
        except Exception as e:
            print(f"⚠️  Warning: Could not calculate hours: {e}")
            return 0.0
    
    def process_time_entries(self, entries: List[Dict], start_date: datetime, end_date: datetime) -> List[Dict]:
        """Procesa los registros de Clockify y los convierte al formato Excel"""
        processed = []
        aggregated = {}
        
        for entry in entries:
            time_interval = entry.get('timeInterval', {})
            start_str = time_interval.get('start')
            
            if not start_str:
                continue
            
            try:
                entry_date = datetime.fromisoformat(start_str.replace('Z', '+00:00'))
                date_only = entry_date.date()
                
                description = entry.get('description', 'Sin descripción')
                
                hours = self.calculate_hours(time_interval)
                jira_id, cleaned_description = self.search_jira_issue_id(description)
                work_location = self.get_work_location(entry_date)
                
                row_date = entry_date.strftime('%m/%d/%Y')
                normalized_description = ' '.join(cleaned_description.split()).lower()
                key = (row_date, jira_id, normalized_description)
                if key not in aggregated:
                    aggregated[key] = {
                        'Subproyecto': 'INGEA - (23006.01) IngeaAPP',
                        'Fecha': row_date,
                        'Horas': hours,
                        'Tipo': 'Técnico',
                        'Etapa': '00-Otro',
                        'Area': '00-Otro',
                        'Lugar Trabajo': work_location,
                        'ID Tarea (Planner/Jira)': jira_id,
                        'Comentario': cleaned_description
                    }
                else:
                    aggregated[key]['Horas'] += hours
            except Exception as e:
                print(f"⚠️  Warning: Error processing entry: {e}")
                continue

        processed.extend(aggregated.values())
        
        # Agrega el registro de Descanso para cada día de lunes a viernes en el rango
        current_date = start_date.date()
        end_date_only = end_date.date()
        
        while current_date <= end_date_only:
            # Solo agregar descanso si es día de semana (lunes=0 a viernes=4)
            current_datetime = datetime.combine(current_date, datetime.min.time())
            if current_datetime.weekday() < 5:  # 0-4 son lunes a viernes
                processed.append({
                    'Subproyecto': 'INGEA - (23006.01) IngeaAPP',
                    'Fecha': current_date.strftime('%m/%d/%Y'),
                    'Horas': 0.5,
                    'Tipo': 'Descanso',
                    'Etapa': '00-Otro',
                    'Area': '00-Otro',
                    'Lugar Trabajo': self.get_work_location(current_datetime),
                    'ID Tarea (Planner/Jira)': 'N/A',
                    'Comentario': 'N/A'
                })
            current_date += timedelta(days=1)
        
        # Ordena por fecha
        processed.sort(key=lambda x: datetime.strptime(x['Fecha'], '%m/%d/%Y'))
        
        return processed
    
    def create_excel(self, data: List[Dict], output_file: str):
        """Crea el archivo Excel con el formato especificado"""
        # Crear carpeta reportes si no existe
        reportes_dir = os.path.join(os.path.dirname(__file__), 'reportes')
        os.makedirs(reportes_dir, exist_ok=True)
        
        # Si output_file es solo el nombre (sin path), guardarlo en reportes/
        if not os.path.dirname(output_file):
            output_file = os.path.join(reportes_dir, output_file)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte Horas"
        
        # Define headers
        headers = [
            'Subproyecto',
            'Fecha',
            'Horas',
            'Tipo',
            'Etapa',
            'Area',
            'Lugar Trabajo',
            'ID Tarea (Planner/Jira)',
            'Comentario'
        ]
        
        # Style for headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Write data
        for row_num, entry in enumerate(data, 2):
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = entry.get(header, '')
                cell.border = border
                
                # Alignment
                if header in ['Horas']:
                    cell.alignment = Alignment(horizontal="right")
                elif header == 'Fecha':
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="left")
        
        # Adjust column widths
        column_widths = {
            'A': 20,  # Subproyecto
            'B': 12,  # Fecha
            'C': 8,   # Horas
            'D': 12,  # Tipo
            'E': 12,  # Etapa
            'F': 12,  # Area
            'G': 15,  # Lugar Trabajo
            'H': 20,  # ID Tarea
            'I': 50   # Comentario
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Save workbook
        wb.save(output_file)
        print(f"✅ Excel file created: {os.path.abspath(output_file)}")
    
    def generate_report(self, start_date: datetime, end_date: datetime, output_file: str):
        """Genera el reporte completo desde Clockify a Excel"""
        print(f"📊 Generating report from {start_date.date()} to {end_date.date()}...")
        
        # Fetch time entries
        print("⏳ Fetching Clockify entries...")
        entries = self.fetch_time_entries(start_date, end_date)
        print(f"✅ Found {len(entries)} time entries")
        
        # Process entries
        print("⚙️  Processing entries...")
        processed_data = self.process_time_entries(entries, start_date, end_date)
        print(f"✅ Processed {len(processed_data)} records")
        
        # Create Excel
        print("📝 Creating Excel file...")
        self.create_excel(processed_data, output_file)
        
        print(f"✨ Report generated successfully!")


def main():
    """Función principal para generar el reporte"""
    converter = ClockifyToExcelConverter()
    
    # Ejemplo: Reporte de la semana anterior
    today = datetime.now()
    start_of_previous_week = today - timedelta(days=today.weekday() + 7)
    # AGREGAR QUE CUALQUIER DIA DE LA SEMANA ANTERIOR QUE ESTE EN EL MES ANTERIOR, IGNORARLO
    end_of_previous_week = start_of_previous_week + timedelta(days=6)
    
    output_file = f"reporte_horas_{start_of_previous_week.strftime('%Y%m%d')}_to_{end_of_previous_week.strftime('%Y%m%d')}.xlsx"
    
    converter.generate_report(start_of_previous_week, end_of_previous_week, output_file)

if __name__ == "__main__":
    main()
